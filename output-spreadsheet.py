#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re

from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl import load_workbook


def iscontain(cell, target):
    if cell.value is None:
        return False

    TARGET_VALUE_PATTERN = re.compile(target)
    if TARGET_VALUE_PATTERN.match(cell.value):
        return True

    return False

def isempty(sheet):
    if sheet.max_row == 1 and sheet.max_column == 1 \
            and sheet['A1'].value is None:
        return True

    return False


def matchCells(sheet, target):
    cells = []

    if isempty(sheet):
        return cells

    if sheet.max_row > 1:
        for row in sheet:
            for cell in row:
                if iscontain(cell, target):
                    cells.append(cell)
    
    return cells


def clipCells(sheet, target_cell, row_width=1, column_width=1):

    coord_r, coord_c = target_cell.row, target_cell.column
    coord_cell = sheet.cell(row=coord_r, column=coord_c)

    if coord_cell != target_cell:
        print("if coord_cell != target_cell: result is false")
        assert(False)
        return False

    range_selected = []
    for i in range(coord_r, coord_r + row_width):
        row_selected = []
        for j in range(coord_c, coord_c + column_width):
            cell = sheet.cell(row=i, column=j)
            row_selected.append(cell)

        range_selected.append(row_selected)

    return range_selected

# read_filename = 'books.odp' #! not support
read_filename = 'books.xlsx'
output_filename = 'books2.xlsx'

# wb = Workbook()
wb = load_workbook(read_filename)

#dest_filename = 'empty_book.xlsx'


sheets = wb.worksheets
# sheets = wb.sheetnames

for sheet in sheets:
    print(sheet)
    print(type(sheet))
    
    if sheet.max_row > 1:
        for row in sheet:
            for cell in row:
                if iscontain(cell, "statistics"):
                    selected_cells = clipCells(sheet, cell, 3, 2)
                    print(selected_cells)
                    for r in selected_cells:
                        for c in r:
                            print(c.value)

# for row in ws.iter_rows(min_row=2):
# for column in ws.iter_columns(min_column=2):

cells = []
# sheet = sheets[0]
# sheet = sheets.get_sheet_by_name('books')
sheet = wb.get_sheet_by_name('books')

#cells = matchCells(sheet, "Effective")
cells = matchCells(sheet, "機械学習")
for c in cells:
    cells = clipCells(sheet, c, row_width=3, column_width=2)
    print(c.value)
# for sheet in sheets:
#     cells = matchCells(sheet, "Effective")
#     print(len(cells))
#     print(cells)
#     for cell in cells:
#         print("{}[{},{}]".format(cell.value, cell.row, cell.column))

# sheet = wb.active
ws1 = wb.create_sheet('destination')
print(ws1.title)

print(type(cells))

for i, row in enumerate(cells, start=1):
    for j, cell in enumerate(row, start=1):
        print("{},{}, cell={}".format(i, j, cell.value))
        ws1.cell(row=i, column=j, value=cell.value)


# for i, cell in enumerate(cells, start=1):
#     print("{}.{} {}".format(i, type(cell), cell.value))
#     print("({0.row},{0.column})={0.value}".format(cell))
#     ws1.cell(row=i, column=1, value=cell.value)
#     # ws1.cell(row=cell.row, column=cell.column, value=cell.value)
#     #print("{}".format(dir(row)))
#     # ws1.cell(row=cell.row,column=cell.column,value=cell.value)
#     # for j, cell in enumerate(row):
#     #     print("{}] = {}".format(j, cell))
#     #     # ws1.cell(row=i, column=j, value=cell.value)
#     #     ws1.cell(row=i, column=j).value = 2
#     #     # ws1.cell(*cell)
#     #     # ws1['A1'] = cell
#     #     # ws1['A1'] = i
#     #     print("cell = {}".format(cell))

#wb.save(read_filename)
#wb.save(read_filename)
wb.save(output_filename)

# ws1 = wb.active
# ws1.title = "range names"
# ws1.sheet_properties.tabColor = "1072BA"

#for row in range(1, 40):
#    ws1.append(range(600))

# [print(elem) for elem in dir(ws1.columns)]
# [print(elem) for elem in dir(ws1.rows)]

assert(False)
ws2 = wb.copy_worksheet(ws1)
ws2.title = "copied sheet"
#ws2.copy_worksheet(ws1)

#ws2 = wb.create_sheet(title="Pi")
#for row in range(1, 40):
#    ws2.append(range(3))
#
#c = ws2['A2']
#
#print(c)
#
#ws2['A4'] = 42


f5 = ws2['F5']

ft = Font(color=colors.RED)
f5.font = ft
#f5.font.italic = True


#ws3 = wb.create_sheet(title="Date")
#for row in range(10, 20):
#    for col in range(27, 54):
#        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#
#print(ws3['A10'].value)

#wb.save(filename=dest_filename)
